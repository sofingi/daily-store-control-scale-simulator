from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional
import io
import pandas as pd
from openpyxl import load_workbook


def _to_num(v):
    if v is None or v == '': return 0.0
    try: return float(v)
    except Exception:
        s=str(v).replace('Rp','').replace('.','').replace(',','.').strip()
        try:return float(s)
        except:return 0.0


def parse_shopee_orders(file_bytes: bytes) -> pd.DataFrame:
    df=pd.read_excel(io.BytesIO(file_bytes),sheet_name='orders',dtype=object)
    need=['No. Pesanan','Status Pesanan','Waktu Pesanan Dibuat','Nomor Referensi SKU','Nama Produk','Jumlah','Returned quantity','Harga Setelah Diskon','Subtotal Pesanan','Total Pembayaran']
    miss=[c for c in need if c not in df.columns]
    if miss: raise ValueError(f'Kolom Order tidak lengkap: {miss}')
    out=pd.DataFrame({
        'order_id':df['No. Pesanan'].astype(str).str.strip(),
        'order_status':df['Status Pesanan'].astype(str).str.strip(),
        'order_date':pd.to_datetime(df['Waktu Pesanan Dibuat'],errors='coerce').dt.date,
        'sku':df['Nomor Referensi SKU'].fillna('').astype(str).str.strip(),
        'product_name':df['Nama Produk'].fillna('').astype(str),
        'qty':pd.to_numeric(df['Jumlah'],errors='coerce').fillna(0),
        'returned_qty':pd.to_numeric(df['Returned quantity'],errors='coerce').fillna(0),
        'unit_price_export':pd.to_numeric(df['Harga Setelah Diskon'],errors='coerce').fillna(0),
        'subtotal_export':pd.to_numeric(df['Subtotal Pesanan'],errors='coerce').fillna(0),
        'total_payment_export':pd.to_numeric(df['Total Pembayaran'],errors='coerce').fillna(0),
    })
    # Shopee Order export for this store represents product prices in thousands of IDR.
    # Detect safely rather than hard-code: median non-zero unit price < 1000 => multiply by 1000.
    nz=out.loc[out.unit_price_export>0,'unit_price_export']
    mult=1000.0 if len(nz) and nz.median()<1000 else 1.0
    out['unit_price_idr']=out.unit_price_export*mult
    out['line_sales_idr']=out.unit_price_idr*out.qty
    out['net_qty']=(out.qty-out.returned_qty).clip(lower=0)
    out=out.dropna(subset=['order_date'])
    return out


def parse_bigseller_hpp_master(file_bytes: bytes) -> pd.DataFrame:
    xls=pd.ExcelFile(io.BytesIO(file_bytes))
    if 'SKU' not in xls.sheet_names: raise ValueError('Sheet SKU BigSeller tidak ditemukan')
    df=pd.read_excel(io.BytesIO(file_bytes),sheet_name='SKU',dtype=object)
    if 'Nomor SKU' not in df or 'Modal Referensi' not in df: raise ValueError('Kolom Nomor SKU / Modal Referensi tidak ditemukan')
    out=pd.DataFrame({
        'sku':df['Nomor SKU'].astype(str).str.strip(),
        'product_title':df.get('Judul',pd.Series(index=df.index,dtype=object)).fillna('').astype(str),
        'unit_hpp':pd.to_numeric(df['Modal Referensi'],errors='coerce'),
        'sku_type':df.get('Jenis SKU',pd.Series(index=df.index,dtype=object)).fillna('').astype(str),
    })
    return out.drop_duplicates('sku',keep='last')


def parse_bigseller_sku_profit_hpp(file_bytes: bytes) -> pd.DataFrame:
    """Secondary HPP source from BigSeller Keuntungan SKU Gudang.
    Derives weighted unit HPP = Total Modal / Jumlah Produk yang Terjual.
    This is only a fallback when the master SKU snapshot does not contain the SKU.
    """
    df=pd.read_excel(io.BytesIO(file_bytes),header=1,dtype=object)
    need=['SKU Gudang','Total Modal','Jumlah Produk yang Terjual']
    miss=[c for c in need if c not in df.columns]
    if miss: raise ValueError(f'Kolom Keuntungan SKU Gudang tidak lengkap: {miss}')
    out=pd.DataFrame({
        'sku':df['SKU Gudang'].fillna('').astype(str).str.strip(),
        'product_title':df.get('Nama SKU Gudang',pd.Series(index=df.index,dtype=object)).fillna('').astype(str),
        'total_hpp':pd.to_numeric(df['Total Modal'],errors='coerce'),
        'units_sold':pd.to_numeric(df['Jumlah Produk yang Terjual'],errors='coerce'),
    })
    out=out[(out.sku!='') & out.total_hpp.notna() & out.units_sold.gt(0)].copy()
    out['unit_hpp']=out.total_hpp/out.units_sold
    out['sku_type']='BIGSELLER_SKU_PROFIT_FALLBACK'
    return out[['sku','product_title','unit_hpp','sku_type']].drop_duplicates('sku',keep='last')


def merge_hpp_sources(primary: pd.DataFrame, fallback: Optional[pd.DataFrame]=None) -> pd.DataFrame:
    """Primary master wins. Fallback only fills SKUs absent from primary."""
    p=primary.copy() if primary is not None else pd.DataFrame(columns=['sku','product_title','unit_hpp','sku_type'])
    if fallback is None or fallback.empty:
        return p.drop_duplicates('sku',keep='last')
    f=fallback.copy()
    f=f[~f.sku.isin(set(p.sku.astype(str)))]
    return pd.concat([p,f],ignore_index=True).drop_duplicates('sku',keep='first')


def hpp_coverage_audit(order_lines: pd.DataFrame, hpp_master: pd.DataFrame) -> dict:
    refs=order_lines.copy()
    refs['sku']=refs['sku'].fillna('').astype(str).str.strip()
    master=set(hpp_master['sku'].fillna('').astype(str).str.strip())
    refs['hpp_known']=refs['sku'].isin(master) & refs['sku'].ne('')
    cancelled=refs['order_status'].astype(str).str.lower().eq('batal')
    refs['realized_qty']=(refs['qty']-refs['returned_qty']).clip(lower=0)
    refs.loc[cancelled,'realized_qty']=0
    active=refs[refs.realized_qty>0].copy()
    total_qty=float(active.realized_qty.sum())
    known_qty=float(active.loc[active.hpp_known,'realized_qty'].sum())
    unmatched=(active[~active.hpp_known]
               .groupby(['sku','product_name'],dropna=False,as_index=False)
               .agg(realized_qty=('realized_qty','sum'),order_lines=('order_id','size'))
               .sort_values(['realized_qty','order_lines'],ascending=False))
    return {
        'active_lines':int(len(active)),
        'matched_active_lines':int(active.hpp_known.sum()),
        'line_coverage':float(active.hpp_known.mean()) if len(active) else 1.0,
        'realized_qty':total_qty,
        'matched_realized_qty':known_qty,
        'qty_coverage':float(known_qty/total_qty) if total_qty else 1.0,
        'unmatched_unique_skus':int(unmatched.sku.nunique()),
        'unmatched':unmatched,
    }


def _income_sheet_rows(file_bytes: bytes, sheet_name: str) -> Iterable[tuple]:
    wb=load_workbook(io.BytesIO(file_bytes),read_only=True,data_only=True)
    ws=wb[sheet_name]
    # Shopee exports have dimension ref=A1 even when the sheet contains millions of cells.
    ws.reset_dimensions()
    yield from ws.iter_rows(values_only=True)


def parse_income_detail(file_bytes: bytes) -> pd.DataFrame:
    """Fast targeted parser for Shopee Income detail.
    Shopee's workbook writes worksheet dimension=A1 even for huge sheets. We only
    need order id, order-created date, release date and Total Penghasilan for V1.6,
    so parsing raw XLSX XML is much faster than materializing all 61 columns.
    """
    import zipfile, re
    from lxml import etree
    z=zipfile.ZipFile(io.BytesIO(file_bytes))
    # shared strings
    ns='{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    shared=[]
    if 'xl/sharedStrings.xml' in z.namelist():
        for _,el in etree.iterparse(z.open('xl/sharedStrings.xml'),events=('end',),tag=ns+'si'):
            shared.append(''.join(el.itertext())); el.clear()
    # workbook sheet -> relationship id
    wb=etree.parse(z.open('xl/workbook.xml'))
    rels=etree.parse(z.open('xl/_rels/workbook.xml.rels'))
    rid_to_target={e.get('Id'):e.get('Target').lstrip('/') for e in rels.getroot()}
    sheet_paths=[]
    rns='{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    for sh in wb.findall('.//'+ns+'sheet'):
        name=sh.get('name','')
        if name == 'Penghasilan' or name.startswith('Penghasilan -'):
            target=rid_to_target.get(sh.get(rns))
            if target:
                if not target.startswith('xl/'): target='xl/'+target
                sheet_paths.append((name,target))
    target_cols={'C':'order_id','G':'order_date_income','H':'release_date','K':'Total Penghasilan'}
    records=[]
    def value_of(c):
        v=c.find(ns+'v')
        if v is None:return None
        txt=v.text
        if c.get('t')=='s':
            try:return shared[int(txt)]
            except:return txt
        return txt
    for sheet_name,path in sheet_paths:
        current={}
        current_row=None
        for event,el in etree.iterparse(z.open(path),events=('end',),tag=(ns+'c',ns+'row')):
            if el.tag==ns+'c':
                ref=el.get('r',''); m=re.match(r'([A-Z]+)(\d+)',ref)
                if m and m.group(1) in target_cols and int(m.group(2))>=4:
                    current[target_cols[m.group(1)]]=value_of(el)
                el.clear()
            else:
                rnum=int(el.get('r','0'))
                if rnum>=4 and current.get('order_id'):
                    rec={'order_id':str(current.get('order_id')).strip(),'income_sheet':sheet_name,
                         'order_date_income':pd.to_datetime(current.get('order_date_income'),errors='coerce').date() if current.get('order_date_income') else None,
                         'release_date':pd.to_datetime(current.get('release_date'),errors='coerce').date() if current.get('release_date') else None,
                         'Total Penghasilan':pd.to_numeric(current.get('Total Penghasilan'),errors='coerce')}
                    records.append(rec)
                current={}; el.clear()
    return pd.DataFrame(records)

def parse_income_seller_fee(file_bytes: bytes) -> pd.DataFrame:
    df=pd.read_excel(io.BytesIO(file_bytes),sheet_name='Seller Fee',header=1,dtype=object)
    cols=['Biaya Platform','Biaya Gratis Ongkir XTRA','Biaya Layanan','Biaya Promosi','Biaya Lainnya']
    out=pd.DataFrame({'order_id':df['No. Pesanan'].astype(str).str.strip()})
    for c in cols: out[c]=pd.to_numeric(df.get(c,0),errors='coerce').fillna(0)
    return out.groupby('order_id',as_index=False)[cols].sum()


def attach_hpp(order_lines: pd.DataFrame, hpp_master: pd.DataFrame) -> pd.DataFrame:
    out=order_lines.merge(hpp_master[['sku','unit_hpp']],on='sku',how='left')
    out['line_hpp']=out.unit_hpp*out.net_qty
    out['hpp_known']=out.unit_hpp.notna()
    return out


def order_level_economics(order_lines_hpp: pd.DataFrame, income_details: list[pd.DataFrame]) -> pd.DataFrame:
    o=order_lines_hpp.copy()
    agg=o.groupby('order_id',as_index=False).agg(
        order_date=('order_date','first'), order_status=('order_status','first'),
        product_sales=('line_sales_idr','sum'), hpp=('line_hpp','sum'),
        order_qty=('net_qty','sum'), sku_lines=('sku','size'),
        hpp_covered_lines=('hpp_known','sum'),
    )
    agg['hpp_coverage']=agg.hpp_covered_lines/agg.sku_lines.replace(0,pd.NA)
    cancelled=agg.order_status.astype(str).str.lower().eq('batal')
    agg.loc[cancelled,'hpp']=0.0
    inc=pd.concat([x for x in income_details if x is not None and not x.empty],ignore_index=True) if any(x is not None and not x.empty for x in income_details) else pd.DataFrame()
    if inc.empty:
        agg['release_date']=pd.NaT; agg['total_income']=pd.NA; agg['profit_status']='ESTIMATED'
        return agg
    # One order can appear as multiple product rows. For exact settlement amount, use max absolute Total Penghasilan
    # when duplicated; this matches Shopee's order-level release total in current exports.
    inc['abs_total']=inc['Total Penghasilan'].abs() if 'Total Penghasilan' in inc else 0
    core=inc.sort_values('abs_total').groupby('order_id',as_index=False).tail(1)
    keep=['order_id','release_date','Total Penghasilan']
    core=core[[c for c in keep if c in core]].rename(columns={'Total Penghasilan':'total_income'})
    agg=agg.merge(core,on='order_id',how='left')
    agg['profit_status']=agg.total_income.notna().map({True:'FINAL',False:'ESTIMATED'})
    agg.loc[agg.order_status.str.lower().eq('batal'),'profit_status']='FINAL_CANCELLED'
    return agg


def fee_rate_benchmark(order_econ: pd.DataFrame) -> float:
    # Use realized income / product sales to infer total marketplace deductions when detail exists.
    x=order_econ[(order_econ.profit_status=='FINAL') & (order_econ.product_sales>0) & order_econ.total_income.notna()]
    if x.empty:return 0.24
    deduction=1-(x.total_income.sum()/x.product_sales.sum())
    return float(min(max(deduction,0),0.60))


def daily_profit_from_orders(order_econ: pd.DataFrame, ads_daily: Optional[pd.DataFrame]=None, estimated_fee_rate: Optional[float]=None) -> pd.DataFrame:
    d=order_econ.copy()
    rate=fee_rate_benchmark(d) if estimated_fee_rate is None else estimated_fee_rate
    d['financial_income']=d.total_income
    estmask=d.profit_status.eq('ESTIMATED')
    d.loc[estmask,'financial_income']=d.loc[estmask,'product_sales']*(1-rate)
    d.loc[d.profit_status.eq('FINAL_CANCELLED'),'financial_income']=0
    d['profit_before_ads']=d.financial_income-d.hpp.fillna(0)
    d['hpp_incomplete']=d.hpp_coverage.lt(1)
    daily=d.groupby('order_date',as_index=False).agg(
        order_sales=('product_sales','sum'), financial_income=('financial_income','sum'), hpp=('hpp','sum'),
        profit_before_ads=('profit_before_ads','sum'), orders=('order_id','nunique'),
        final_orders=('profit_status',lambda s:(s=='FINAL').sum()), estimated_orders=('profit_status',lambda s:(s=='ESTIMATED').sum()),
        incomplete_hpp_orders=('hpp_incomplete','sum'))
    daily['settlement_coverage']=daily.final_orders/daily.orders.replace(0,pd.NA)
    daily['hpp_coverage']=1-daily.incomplete_hpp_orders/daily.orders.replace(0,pd.NA)
    daily['estimated_fee_rate']=rate
    daily['ads_spend']=pd.NA
    if ads_daily is not None and not ads_daily.empty:
        a=ads_daily.copy(); a['order_date']=pd.to_datetime(a['metric_date']).dt.date
        a=a.groupby('order_date',as_index=False).ads_spend.sum()
        daily=daily.merge(a,on='order_date',how='left',suffixes=('','_a'))
        if 'ads_spend_a' in daily:
            daily['ads_spend']=daily.pop('ads_spend_a')
    daily['control_profit']=daily.profit_before_ads-pd.to_numeric(daily.ads_spend,errors='coerce').fillna(0)
    daily['control_margin']=daily.control_profit/daily.order_sales.replace(0,pd.NA)
    daily['profit_status']='ESTIMATED'
    daily.loc[(daily.settlement_coverage>=0.98)&(daily.hpp_coverage>=0.98)&daily.ads_spend.notna(),'profit_status']='FINAL'
    daily.loc[(daily.settlement_coverage<0.70)|(daily.hpp_coverage<0.90),'profit_status']='PARTIAL'
    return daily


def period_profit(order_econ: pd.DataFrame, total_ads_spend: float=0.0) -> dict:
    x=order_econ.copy()
    rate=fee_rate_benchmark(x)
    x['financial_income']=x.total_income
    est=x.profit_status.eq('ESTIMATED')
    x.loc[est,'financial_income']=x.loc[est,'product_sales']*(1-rate)
    x.loc[x.profit_status.eq('FINAL_CANCELLED'),'financial_income']=0
    hpp=x.hpp.fillna(0).sum(); income=x.financial_income.fillna(0).sum(); sales=x.product_sales.sum()
    before=income-hpp; control=before-total_ads_spend
    return {
        'order_sales':float(sales),'financial_income':float(income),'hpp':float(hpp),
        'profit_before_ads':float(before),'ads_spend':float(total_ads_spend),'control_profit':float(control),
        'control_margin':float(control/sales) if sales else None,
        'settlement_coverage':float((x.profit_status=='FINAL').sum()/len(x)) if len(x) else 0,
        'hpp_coverage':float((x.hpp_coverage>=1).sum()/len(x)) if len(x) else 0,
        'estimated_fee_rate':rate,
    }
